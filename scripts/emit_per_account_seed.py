#!/usr/bin/env python3
"""
Per-account SQL seed emitter for NuRock legacy workbooks.

Reads each property workbook from PROJECT_FILES_DIR, runs the account-detail
parsers over the Water / Hse Meters / Garbage / Phone&Cable sheets, and emits
a SQL file that inserts:

    * Real vendors (one per vendor_name found)
    * Real utility_accounts (one per account_number per GL per property)
    * Real invoices (one per account per month with the actual dollar amount)

This replaces the synthetic `HIST-S-<code>-<gl>-<desc>-<year>-<mm>` rollup
invoices from the old importer. Summary view totals will still reconcile
because the sum across all per-account invoices = the old rollup totals.

Output SQL is idempotent (uses `on conflict do nothing`) so it can be applied
over an existing database without duplicating data. You should TRUNCATE the
affected tables before re-applying if you want to start clean — see the
companion migration 0012_reset_historical.sql for a safe reset path.
"""

import os
import re
import sys
from decimal import Decimal
from pathlib import Path
from datetime import date

# Ensure script dir is on path
sys.path.insert(0, str(Path(__file__).resolve().parent))
from _account_detail_parser import (  # noqa: E402
    parse_water_accounts, parse_house_meters_accounts, parse_garbage_accounts,
    parse_phone_cable_accounts, parse_fedex_accounts, parse_vacant_units_accounts,
    parse_fixed_accounts,
)

from openpyxl import load_workbook  # noqa: E402
import warnings; warnings.filterwarnings("ignore")


# Map of property CODE → .xlsx path (expects converted .xlsx files)
# Falls back to scanning /tmp/converted/ which is where libreoffice dumps.
PROPERTY_WORKBOOKS = {
    "508": "508Hearthstone.xlsx",
    "509": "509Heritage_at_Walton_Reserve.xlsx",
    "514": "514Hidden_Creste.xlsx",
    "515": "515Tuscany.xlsx",
    "516": "516Heritage_McDonough.xlsx",
    "555": "555Sunset_Pointe.xlsx",
    "558": "558Onion_Creek.xlsx",
    "559": "559Eastland.xlsx",
    "560": "560Heritage_Park_Vista.xlsx",
    "561": "561Stalcup.xlsx",
    "562": "562EC_Tyler.xlsx",
    "601": "601Town_Park_Crossing.xlsx",
    "602": "602Vista_Grand.xlsx",
    "603": "603Crystal_Lakes.xlsx",
    "604": "604Heritage_at_Pompano.xlsx",
    "606": "606Haverhill.xlsx",
    "607": "607Residences_at_Marathon_Key.xlsx",
    "608": "608Residences_at_Crystal_Cove.xlsx",
    "610": "610_Residences_at_Naranja_Lakes.xlsx",
    "611": "611_Residences_at_Beverly_Park.xlsx",
}


def sql_str(v) -> str:
    """SQL literal for a string. Handles nulls and single-quote escaping."""
    if v is None:
        return "null"
    return "'" + str(v).replace("'", "''") + "'"


def sql_dec(v) -> str:
    if v is None:
        return "null"
    return f"{float(v):.2f}"


def month_date(year, month):
    """Return a SQL date literal for the first of the month."""
    return f"'{year:04d}-{month:02d}-01'"


# ----------------------------------------------------------------------------
# Data collection
# ----------------------------------------------------------------------------

def collect_all_workbooks(converted_dir):
    """Walk all property workbooks and build a unified record set."""
    all_water         = {}
    all_house_meters  = {}
    all_garbage       = {}
    all_phone_cable   = {}
    all_fedex         = {}
    all_vacant        = {}
    all_fixed         = {}

    for code, fname in PROPERTY_WORKBOOKS.items():
        path = Path(converted_dir) / fname
        if not path.exists():
            print(f"  [skip] {code} — no workbook at {path}", file=sys.stderr)
            continue

        try:
            wb = load_workbook(path, data_only=True)
        except Exception as e:
            print(f"  [fail] {code} — {e}", file=sys.stderr)
            continue

        year = date.today().year

        if "Water" in wb.sheetnames:
            r = parse_water_accounts(wb["Water"], year)
            if r and r.get("accounts"):
                all_water[code] = r

        for hm_sheet in ("Hse Meters", "House Meters", "HseMeters"):
            if hm_sheet in wb.sheetnames:
                r = parse_house_meters_accounts(wb[hm_sheet], year)
                if r and r.get("accounts"):
                    all_house_meters[code] = r
                break

        if "Garbage" in wb.sheetnames:
            r = parse_garbage_accounts(wb["Garbage"], year)
            if r and r.get("accounts"):
                all_garbage[code] = r

        for pc_sheet in ("Phone&Cable", "Phone & Cable", "Phone Cable"):
            if pc_sheet in wb.sheetnames:
                r = parse_phone_cable_accounts(wb[pc_sheet], year)
                if r and r.get("accounts"):
                    all_phone_cable[code] = r
                break

        if "FedEx" in wb.sheetnames:
            r = parse_fedex_accounts(wb["FedEx"], year)
            if r and r.get("accounts"):
                all_fedex[code] = r

        for vu_sheet in ("Vac Units", "Vacant Units", "VacUnits"):
            if vu_sheet in wb.sheetnames:
                r = parse_vacant_units_accounts(wb[vu_sheet], year)
                if r and r.get("accounts"):
                    all_vacant[code] = r
                break

        if "FIXED" in wb.sheetnames:
            r = parse_fixed_accounts(wb["FIXED"], year)
            if r and r.get("accounts"):
                all_fixed[code] = r

    return all_water, all_house_meters, all_garbage, all_phone_cable, all_fedex, all_vacant, all_fixed


# ----------------------------------------------------------------------------
# SQL emission
# ----------------------------------------------------------------------------

def emit_reset_header():
    return """-- ============================================================================
-- 0012 — Historical per-account seed (replaces 0005 summary rollups)
--
-- Deletes old synthetic HIST-S-<code>-<gl>-* rollup invoices and inserts
-- real per-account utility_accounts + per-account-per-month invoices parsed
-- from the legacy property workbooks.
--
-- Safe to apply idempotently. Clears only rows with source_reference in the
-- list below; any live/extracted data is preserved.
--
-- Requires migrations 0001, 0006 applied first. Skips invoice_line_items
-- cleanup gracefully if the table doesn't exist (to let a partial schema
-- still apply this seed).
-- ============================================================================

do $$
begin
    if exists (select 1 from information_schema.tables
               where table_schema = 'public' and table_name = 'invoice_line_items') then
        execute $sql$
            delete from invoice_line_items
            where invoice_id in (
                select id from invoices
                where source_reference in (
                    'historical_import_summary',
                    'historical_import_water_usage',
                    'historical_import_per_account'
                )
            )
        $sql$;
    end if;
end $$;

delete from invoices
where source_reference in (
    'historical_import_summary',
    'historical_import_water_usage',
    'historical_import_per_account'
);

delete from utility_accounts
where account_number like 'HIST-%'
   or account_number like 'PERACCT-%';


-- Ensure a unique index on vendors.name so the ON CONFLICT below works.
-- (0001 created vendors with no uniqueness guarantee; we add it here since
--  the per-account seed relies on vendor names being unique identifiers.)
create unique index if not exists vendors_name_key on vendors (name);

"""


def emit_vendors(water, hse_meters, garbage, phone_cable, fedex=None, vacant=None, fixed=None):
    """Emit all distinct vendors discovered across the workbooks."""
    lines = ["-- Vendors discovered from legacy workbooks"]
    seen = set()
    vendor_category = {}
    for src, cat in ((water, "water"),
                     (hse_meters, "electric"),
                     (garbage, "trash"),
                     (phone_cable, "phone"),
                     (fedex or {}, "fedex"),
                     (vacant or {}, "electric"),
                     (fixed or {}, "other")):
        for code, data in src.items():
            # Sheet-level vendor
            name = data.get("vendor_name")
            if name and name not in seen:
                seen.add(name)
                vendor_category[name] = cat
            # Per-row vendors (Phone&Cable typically has these)
            for acct in data.get("accounts", []):
                rv = acct.get("vendor_name")
                if rv and rv not in seen:
                    seen.add(rv)
                    vendor_category[rv] = cat

    if not seen:
        return "\n".join(lines + [""])

    lines.append("insert into vendors (name, category) values")
    rows = [f"  ({sql_str(n)}, {sql_str(vendor_category.get(n,'other'))})"
            for n in sorted(seen)]
    lines.append(",\n".join(rows))
    lines.append("on conflict (name) do nothing;\n")
    return "\n".join(lines)


def emit_utility_accounts(all_data, gl_mapping):
    """
    Emit utility_accounts rows for every (property, account, GL) combination.

    all_data:   {property_code: parser_result}
    gl_mapping: fn(account_category) → GL code. For water, multiple GLs per
                account. For electric/trash/telecom, one GL per sheet.
    """
    lines = []
    rows = []  # (property_code, vendor_name, gl_code, account_number, description, meter_id, active)

    for pc, data in sorted(all_data.items()):
        sheet_vendor = data.get("vendor_name") or "Unknown vendor"
        for acct in data.get("accounts", []):
            acct_num = acct.get("account_number")
            if not acct_num:
                continue
            desc    = acct.get("description") or acct.get("meter_id") or ""
            meter   = acct.get("meter_id")
            # Per-row vendor takes precedence (Phone&Cable). Fall back to sheet vendor.
            vendor_name = acct.get("vendor_name") or sheet_vendor

            if callable(gl_mapping):
                gls = gl_mapping(acct)
            else:
                gls = [gl_mapping]

            for gl in gls:
                rows.append((pc, vendor_name, gl, acct_num, desc, meter))

    if not rows:
        return ""

    lines.append(f"-- {len(rows):,} per-account utility_accounts rows")
    lines.append("with ua(property_code, vendor_name, gl_code, account_number, description, meter_id) as (values")
    value_rows = [
        f"  ({sql_str(pc)}, {sql_str(vn)}, {sql_str(gl)}, "
        f"{sql_str(an)}, {sql_str(desc)}, {sql_str(mid)})"
        for pc, vn, gl, an, desc, mid in rows
    ]
    lines.append(",\n".join(value_rows))
    lines.append(")")
    lines.append("""insert into utility_accounts
  (property_id, vendor_id, gl_account_id, account_number, description, meter_id,
   sub_code, baseline_window_months, variance_threshold_pct, usage_unit, active)
select
  p.id, v.id, g.id, ua.account_number, ua.description, ua.meter_id,
  '00', 12, 3.00, null, true
from ua
  join properties  p on p.code = ua.property_code
  join vendors     v on v.name = ua.vendor_name
  join gl_accounts g on g.code = ua.gl_code
on conflict (vendor_id, account_number) do nothing;
""")
    return "\n".join(lines)


def emit_water_invoices(water):
    """
    Emit one invoice per (account × month × GL) for water.
    Each Water sheet entry has water/sewer/irrigation breakdowns — we emit
    one invoice per category per month, linked to the (account, GL) utility_account.
    """
    if not water:
        return ""
    lines = []
    rows = []

    for pc, data in sorted(water.items()):
        vendor_name = data.get("vendor_name") or ""
        inv_base    = data.get("invoice_number_base") or ""
        year        = data.get("year", date.today().year)

        for acct in data.get("accounts", []):
            an = acct.get("account_number")
            if not an: continue
            for month, md in acct.get("by_month", {}).items():
                if not isinstance(md, dict): continue
                # Three potential GL invoices per month: Water / Sewer / Irrigation
                for role, gl in (("water", "5120"), ("sewer", "5125"), ("irrigation", "5122")):
                    amount = md.get(role)
                    if not amount or amount <= 0: continue
                    invoice_number = f"{inv_base or 'HIST'}-{pc}-{an}-{gl}-{year}-{month:02d}"
                    rows.append({
                        "property_code":  pc,
                        "vendor_name":    vendor_name,
                        "account_number": an,
                        "gl_code":        gl,
                        "invoice_number": invoice_number,
                        "year":           year,
                        "month":          month,
                        "amount":         amount,
                        "days":           md.get("days"),
                        "period":         md.get("period"),
                    })

    if not rows:
        return ""

    lines.append(f"-- {len(rows):,} per-account water invoices (replacing HIST-S- rollups)")
    lines.append("with inv(property_code, vendor_name, account_number, gl_code, invoice_number, yr, mo, amount, days, period) as (values")
    vals = [
        f"  ({sql_str(r['property_code'])}, {sql_str(r['vendor_name'])}, "
        f"{sql_str(r['account_number'])}, {sql_str(r['gl_code'])}, "
        f"{sql_str(r['invoice_number'])}, {r['year']}, {r['month']}, "
        f"{sql_dec(r['amount'])}, "
        f"{r['days'] if r['days'] is not None else 'null'}, "
        f"{sql_str(r['period'])})"
        for r in rows
    ]
    lines.append(",\n".join(vals))
    lines.append(")")
    lines.append("""insert into invoices
  (property_id, vendor_id, utility_account_id, gl_account_id,
   invoice_number, invoice_date, due_date,
   service_period_start, service_period_end, service_days,
   current_charges, total_amount_due,
   status, source, source_reference, sage_posted_at, gl_coding,
   extraction_confidence, requires_human_review)
select
  p.id, ven.id, ua.id, g.id,
  inv.invoice_number,
  make_date(inv.yr, inv.mo, 1),
  (date_trunc('month', make_date(inv.yr, inv.mo, 1)) + interval '28 days')::date,
  make_date(inv.yr, inv.mo, 1),
  (date_trunc('month', make_date(inv.yr, inv.mo, 1)) + interval '1 month - 1 day')::date,
  coalesce(inv.days, extract(day from (date_trunc('month', make_date(inv.yr, inv.mo, 1))
                                       + interval '1 month - 1 day'))::int),
  inv.amount, inv.amount,
  'paid', 'manual', 'historical_import_per_account',
  make_date(inv.yr, inv.mo, 1),
  '500-' || p.code || '-' || inv.gl_code || '.00',
  1.00, false
from inv
  join properties p    on p.code = inv.property_code
  join vendors    ven  on ven.name = inv.vendor_name
  join gl_accounts g   on g.code = inv.gl_code
  join utility_accounts ua on ua.vendor_id = ven.id
                          and ua.account_number = inv.account_number
                          and ua.gl_account_id = g.id
;
""")
    return "\n".join(lines)


def emit_vacant_charges(vacant):
    """
    Emit per-unit-per-month rows into the dedicated vacant_unit_charges table.

    Vacant electric charges are stored separately from invoices because they
    represent allocated cost off a master vacant-unit account, not real AP
    invoices. The Vacant Units detail page reads from this table directly.
    """
    if not vacant:
        return ""

    rows = []
    for pc, data in sorted(vacant.items()):
        year = data.get("year", date.today().year)
        for acct in data.get("accounts", []):
            unit_number = acct.get("unit_number") or acct.get("description")
            if not unit_number:
                continue
            account_number = acct.get("account_number")
            meter_id = acct.get("meter_id")
            for month, md in acct.get("by_month", {}).items():
                amount = md.get("amount") if isinstance(md, dict) else md
                if not amount or amount <= 0:
                    continue
                rows.append({
                    "property_code":  pc,
                    "unit_number":    str(unit_number),
                    "year":           year,
                    "month":          month,
                    "amount":         amount,
                    "meter_id":       meter_id,
                    "account_number": account_number,
                })

    if not rows:
        return ""

    lines = [f"-- {len(rows):,} vacant unit charges"]
    lines.append("delete from vacant_unit_charges where source = 'historical_import_per_account';")
    lines.append("with vc(property_code, unit_number, yr, mo, amount, meter_id, account_number) as (values")
    vals = [
        f"  ({sql_str(r['property_code'])}, {sql_str(r['unit_number'])}, "
        f"{r['year']}, {r['month']}, {sql_dec(r['amount'])}, "
        f"{sql_str(r['meter_id'])}, {sql_str(r['account_number'])})"
        for r in rows
    ]
    lines.append(",\n".join(vals))
    lines.append(")")
    lines.append("""insert into vacant_unit_charges
  (property_id, unit_number, year, month, amount, meter_id, account_number,
   gl_account_id, gl_coding, source)
select
  p.id, vc.unit_number, vc.yr, vc.mo, vc.amount, vc.meter_id, vc.account_number,
  g.id, '500-' || p.code || '-5114.00',
  'historical_import_per_account'
from vc
  join properties p   on p.code = vc.property_code
  left join gl_accounts g on g.code = '5114'
on conflict (property_id, unit_number, year, month) do update
  set amount = excluded.amount,
      meter_id = excluded.meter_id,
      account_number = excluded.account_number,
      source = excluded.source;
""")
    return "\n".join(lines)


def emit_simple_invoices(all_data, gl_code, source_ref_tag):
    """
    Emit one invoice per (account × month) for Hse Meters / Garbage / Phone&Cable / FedEx.
    Used for sheets where each account row has one dollar amount per month.
    """
    if not all_data:
        return ""
    lines = []
    rows = []

    for pc, data in sorted(all_data.items()):
        sheet_vendor = data.get("vendor_name") or ""
        inv_base     = data.get("invoice_number_base") or ""
        year         = data.get("year", date.today().year)
        sheet_gl     = data.get("gl_code", gl_code)

        for acct in data.get("accounts", []):
            an = acct.get("account_number")
            if not an: continue
            # Per-row vendor takes precedence (Phone&Cable)
            vendor_name = acct.get("vendor_name") or sheet_vendor
            # Per-row gl_code takes precedence (FIXED)
            row_gl = acct.get("gl_code") or sheet_gl
            for month, md in acct.get("by_month", {}).items():
                amount = md.get("amount") if isinstance(md, dict) else md
                if not amount or amount <= 0: continue
                invoice_number = f"{inv_base or 'HIST'}-{pc}-{an}-{year}-{month:02d}"
                rows.append({
                    "property_code":  pc,
                    "vendor_name":    vendor_name,
                    "account_number": an,
                    "gl_code":        row_gl,
                    "invoice_number": invoice_number,
                    "year":           year,
                    "month":          month,
                    "amount":         amount,
                })

    if not rows:
        return ""

    lines.append(f"-- {len(rows):,} per-account invoices for {source_ref_tag}")
    lines.append(f"with inv(property_code, vendor_name, account_number, gl_code, invoice_number, yr, mo, amount) as (values")
    vals = [
        f"  ({sql_str(r['property_code'])}, {sql_str(r['vendor_name'])}, "
        f"{sql_str(r['account_number'])}, {sql_str(r['gl_code'])}, "
        f"{sql_str(r['invoice_number'])}, {r['year']}, {r['month']}, "
        f"{sql_dec(r['amount'])})"
        for r in rows
    ]
    lines.append(",\n".join(vals))
    lines.append(")")
    lines.append("""insert into invoices
  (property_id, vendor_id, utility_account_id, gl_account_id,
   invoice_number, invoice_date, due_date,
   service_period_start, service_period_end, service_days,
   current_charges, total_amount_due,
   status, source, source_reference, sage_posted_at, gl_coding,
   extraction_confidence, requires_human_review)
select
  p.id, ven.id, ua.id, g.id,
  inv.invoice_number,
  make_date(inv.yr, inv.mo, 1),
  (date_trunc('month', make_date(inv.yr, inv.mo, 1)) + interval '28 days')::date,
  make_date(inv.yr, inv.mo, 1),
  (date_trunc('month', make_date(inv.yr, inv.mo, 1)) + interval '1 month - 1 day')::date,
  extract(day from (date_trunc('month', make_date(inv.yr, inv.mo, 1))
                    + interval '1 month - 1 day'))::int,
  inv.amount, inv.amount,
  'paid', 'manual', 'historical_import_per_account',
  make_date(inv.yr, inv.mo, 1),
  '500-' || p.code || '-' || inv.gl_code || '.00',
  1.00, false
from inv
  join properties p    on p.code = inv.property_code
  join vendors    ven  on ven.name = inv.vendor_name
  join gl_accounts g   on g.code = inv.gl_code
  join utility_accounts ua on ua.vendor_id = ven.id
                          and ua.account_number = inv.account_number
                          and ua.gl_account_id = g.id
;
""")
    return "\n".join(lines)


def main():
    converted_dir = os.environ.get("CONVERTED_DIR", "/tmp/converted")
    out_path = os.environ.get("OUTPUT_PATH",
                              "supabase/migrations/0012_per_account_historical.sql")

    print(f"Reading workbooks from {converted_dir}...")
    water, hm, garbage, phone_cable, fedex, vacant, fixed = collect_all_workbooks(converted_dir)
    print(f"  Water:       {len(water)} properties")
    print(f"  Hse Meters:  {len(hm)} properties")
    print(f"  Garbage:     {len(garbage)} properties")
    print(f"  Phone&Cable: {len(phone_cable)} properties")
    print(f"  FedEx:       {len(fedex)} properties")
    print(f"  Vacant:      {len(vacant)} properties")
    print(f"  FIXED:       {len(fixed)} properties")

    # Water: one account can have 3 GL accounts (water/sewer/irrigation)
    def water_gls(acct):
        needed = set()
        for md in acct.get("by_month", {}).values():
            if not isinstance(md, dict): continue
            if md.get("water"):      needed.add("5120")
            if md.get("sewer"):      needed.add("5125")
            if md.get("irrigation"): needed.add("5122")
        return sorted(needed)

    # FIXED: each account has its own gl_code attached at parse time
    def fixed_gls(acct):
        gl = acct.get("gl_code")
        return [gl] if gl else []

    sections = [
        emit_reset_header(),
        emit_vendors(water, hm, garbage, phone_cable, fedex, vacant, fixed),
        emit_utility_accounts(water,       water_gls),
        emit_utility_accounts(hm,          "5112"),
        emit_utility_accounts(garbage,     "5135"),
        emit_utility_accounts(phone_cable, "5635"),
        emit_utility_accounts(fedex,       "5620"),
        emit_utility_accounts(fixed,       fixed_gls),
        emit_water_invoices(water),
        emit_simple_invoices(hm,          "5112", "House Meters"),
        emit_simple_invoices(garbage,     "5135", "Garbage"),
        emit_simple_invoices(phone_cable, "5635", "Phone&Cable"),
        emit_simple_invoices(fedex,       "5620", "FedEx"),
        emit_simple_invoices(fixed,       None,   "FIXED"),
        emit_vacant_charges(vacant),
    ]

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w") as f:
        f.write("\n\n".join(s for s in sections if s))

    print(f"\nWrote {out_path}")


if __name__ == "__main__":
    main()
