#!/usr/bin/env python3
"""
Historical data importer for NuRock Utilities AP.

Emits a single SQL seed file that loads two kinds of history:

  1. Monthly utility totals from each per-property workbook's Summary sheet.
     Inserted as status='paid' invoices so the v_property_summary view rolls
     them up as actuals immediately, and the tracker page shows them.

  2. Multi-year water meter readings from each property's water-usage
     breakdown file. Inserted as status='paid' water invoices with linked
     usage_readings so the variance engine has real baselines on day one.
     Most breakdown files go back to 2013; 2025 is covered fully.

Matches properties by the 3-digit code at the front of the workbook filename
and by word-match on the breakdown filename. Safe to re-run — idempotency
comes from deterministic invoice_number values and ON CONFLICT DO NOTHING.

Usage:
    python scripts/import-historical.py \\
        --input /path/to/historical/files \\
        --output supabase/seed/0005_historical_data.sql

Requires openpyxl, and LibreOffice on PATH for .xls → .xlsx conversion.
"""

import argparse
import re
import subprocess
import sys
import tempfile
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl required: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# Import the sibling water-detail parser
sys.path.insert(0, str(Path(__file__).parent))
from _water_detail_parser import parse_water_detail   # noqa: E402
from _house_meters_parser import parse_house_meters    # noqa: E402

# -----------------------------------------------------------------------------
# Reference data
# -----------------------------------------------------------------------------

PROPERTY_BY_CODE = {
    "508": "Hearthstone Landing",
    "509": "Heritage at Walton Reserve",
    "514": "Hidden Creste",
    "515": "Tuscany Village",
    "516": "Heritage at McDonough",
    "555": "Sunset Pointe",
    "558": "Onion Creek",
    "559": "Eastland",
    "560": "Heritage Park Vista",
    "561": "Stalcup",
    "562": "EC Tyler",
    "601": "Town Park Crossing",
    "602": "Vista Grand",
    "603": "Crystal Lakes",
    "604": "Heritage at Pompano",
    "606": "Haverhill",
    "607": "Marathon Key",
    "608": "Crystal Cove",
    "610": "Naranja Lakes",
    "611": "Residences at Beverly Park",
}

# Map Summary-sheet description text → canonical GL code.
# Order matters: longer keys matched first.
GL_CODE_MAP = [
    ("vacant unit electric", "5114"),
    ("vacant electric",      "5114"),
    ("cluhouse electric",    "5116"),   # frequent source typo
    ("clubhouse electric",   "5116"),
    ("house electric",       "5112"),
    ("cluhouse - water",     "5120"),
    ("clubhouse - water",    "5120"),
    ("clubhouse water",      "5120"),
    ("storm water",          "5120"),
    ("stome water",          "5120"),   # source typo
    ("envir. protect",       "5120"),
    ("environmental",        "5120"),
    ("irrigation",           "5122"),
    ("cluhouse - sewer",     "5125"),
    ("clubhouse - sewer",    "5125"),
    ("sewer",                "5125"),
    ("water",                "5120"),
    ("gas",                  "5130"),
    ("trash removal",        "5135"),
    ("cable television",     "5140"),
    ("telephone",            "5635"),
    ("phone",                "5635"),
    ("fedex",                "5620"),
]

# -----------------------------------------------------------------------------
# Helpers
# -----------------------------------------------------------------------------

def ensure_xlsx(path: Path, workdir: Path) -> Path | None:
    if path.suffix.lower() == ".xlsx":
        return path
    out = workdir / (path.stem + ".xlsx")
    if out.exists():
        return out
    try:
        subprocess.run(
            ["soffice", "--headless", "--convert-to", "xlsx",
             "--outdir", str(workdir), str(path)],
            check=True, capture_output=True, timeout=60,
        )
        return out if out.exists() else None
    except Exception:
        return None

def property_code_from_filename(name: str) -> str | None:
    m = re.match(r"^(\d{3})[A-Za-z_]", name)
    return m.group(1) if m else None

def infer_gl_code(description: str | None) -> str | None:
    if not description:
        return None
    lo = str(description).lower()
    for key, code in GL_CODE_MAP:
        if key in lo:
            return code
    return None

def to_dec(v) -> Decimal | None:
    if v is None or v == "":
        return None
    try:
        d = Decimal(str(v))
        return d if abs(d) > Decimal("0.001") else None
    except Exception:
        return None

def sql_str(v) -> str:
    if v is None: return "null"
    s = str(v).replace("'", "''")
    return f"'{s}'"

def find_year(ws) -> int | None:
    for r in range(1, 6):
        for c in range(1, 5):
            v = ws.cell(r, c).value
            if isinstance(v, int) and 2000 <= v <= 2099:
                return v
            if isinstance(v, str):
                m = re.search(r"\b(20\d{2})\b", v)
                if m: return int(m.group(1))
    return None

# -----------------------------------------------------------------------------
# Summary sheet parser
# -----------------------------------------------------------------------------

def parse_summary(wb, property_code: str) -> list:
    """Return list of dicts: { year, month, gl_code, description, amount }."""
    sheet = None
    for s in wb.sheetnames:
        if s.lower().strip() == "summary":
            sheet = s; break
    if not sheet:
        for s in wb.sheetnames:
            if "summary" in s.lower() and "yc" not in s.lower():
                sheet = s; break
    if not sheet:
        return []

    ws = wb[sheet]
    year = find_year(ws) or date.today().year

    # Find header row (has "GL AC" in column A)
    header_row = None
    for r in range(1, 12):
        v = ws.cell(r, 1).value
        if v and "gl ac" in str(v).lower():
            header_row = r; break
    if not header_row:
        return []

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        desc = ws.cell(r, 2).value
        if not desc:
            continue
        desc_s = str(desc).strip()
        if desc_s.lower().startswith("total"):
            break

        gl_code = infer_gl_code(desc_s)
        if not gl_code:
            # Try col A for explicit GL code
            a = ws.cell(r, 1).value
            if a and re.match(r"^\d{4}$", str(a).strip()):
                gl_code = str(a).strip()
        if not gl_code:
            continue

        # Monthly values in columns C..N
        for m_idx in range(12):
            amount = to_dec(ws.cell(r, 3 + m_idx).value)
            if amount:
                rows.append({
                    "year":        year,
                    "month":       m_idx + 1,
                    "gl_code":     gl_code,
                    "description": desc_s,
                    "amount":      amount,
                })
    return rows

# -----------------------------------------------------------------------------
# Water usage breakdown parser
# -----------------------------------------------------------------------------

def parse_water_breakdown(wb) -> list:
    """
    Returns a list of dicts with historical water usage:
      { period_start, period_end, days, usage, daily_usage, occupancy, unit }

    Handles four layouts observed in the source files:
      A. 'Service Period' header (Onion Creek, Sunset Pointe, Walton Reserve, etc.)
      B. 'Service dates'  header (Heritage McDonough)
      C. 'Read Date'      header with single read dates (Hidden Creste Consumption History)
      D. 'Bills' sheets are skipped — they duplicate Sheet1 with less detail.
    """
    # Prefer Sheet1 or Consumption History; skip Bills (duplicative)
    candidates = [s for s in wb.sheetnames if s.lower() not in ("bills", "sheet2", "sheet3")]
    if not candidates:
        return []
    sheet_name = candidates[0]
    ws = wb[sheet_name]

    # Find header row and column positions
    header_row = None
    cols = {}
    layout = None  # 'period' or 'readdate'

    for r in range(1, 14):
        for c in range(1, min(ws.max_column + 1, 20)):
            v = ws.cell(r, c).value
            if not isinstance(v, str):
                continue
            lo = v.lower().strip()
            if layout is None:
                if "service period" in lo or "service dates" in lo:
                    layout = "period"; cols["period"] = c; header_row = r
                elif lo == "read date" or "reading date" in lo:
                    layout = "readdate"; cols["read_date"] = c; header_row = r

            if header_row == r:
                if ("days" not in cols) and (lo == "days" or "service days" in lo):
                    cols["days"] = c
                elif ("usage" not in cols) and ("total usage" in lo or "total consumption" in lo or lo == "consumption"):
                    cols["usage"] = c
                elif ("daily" not in cols) and ("daily usage" in lo or "average usage" in lo):
                    cols["daily"] = c
                elif ("occupancy" not in cols) and ("occupancy" in lo or "occupied" in lo):
                    cols["occupancy"] = c

        if header_row:
            break

    if not header_row or layout is None:
        return []

    records = []
    last_read_date = None
    unit = "gallons"  # default
    # Hidden Creste's "Consumption History" sheet uses CCF — detect from header above the column
    if "period" in cols:
        hdr_cell_above = ws.cell(header_row - 1, cols.get("usage", 1)).value
        if hdr_cell_above and "ccf" in str(hdr_cell_above).lower():
            unit = "ccf"
    # Check near the top of the sheet for a CCF marker
    for r in range(1, header_row):
        for c in range(1, min(ws.max_column + 1, 15)):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "ccf" in v.lower():
                unit = "ccf"

    for r in range(header_row + 1, ws.max_row + 1):
        ps = pe = None
        days = None

        if layout == "period":
            period_raw = ws.cell(r, cols["period"]).value
            if not period_raw:
                continue
            ps, pe = parse_period_text(str(period_raw))
            if not ps:
                continue
            if "days" in cols:
                days = to_dec(ws.cell(r, cols["days"]).value)
                days = int(days) if days else None

        elif layout == "readdate":
            rd = ws.cell(r, cols["read_date"]).value
            if not rd:
                continue
            # Parse as date
            if isinstance(rd, datetime):
                cur = rd.date()
            elif isinstance(rd, date):
                cur = rd
            elif isinstance(rd, str):
                m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", rd.strip())
                if not m:
                    continue
                mm, dd, yy = int(m.group(1)), int(m.group(2)), int(m.group(3))
                if yy < 100: yy += 2000
                try:
                    cur = date(yy, mm, dd)
                except Exception:
                    continue
            else:
                continue
            ps = last_read_date or cur
            pe = cur
            last_read_date = cur
            if "days" in cols:
                days = to_dec(ws.cell(r, cols["days"]).value)
                days = int(days) if days else None
            if not days and ps and pe:
                days = (pe - ps).days or None

        usage = to_dec(ws.cell(r, cols["usage"]).value)   if "usage" in cols else None
        daily = to_dec(ws.cell(r, cols["daily"]).value)   if "daily" in cols else None
        occ   = to_dec(ws.cell(r, cols["occupancy"]).value) if "occupancy" in cols else None

        if not usage and not daily:
            continue

        records.append({
            "period_start": ps,
            "period_end":   pe,
            "days":         days,
            "usage":        usage,
            "daily_usage":  daily,
            "occupancy":    occ,
            "unit":         unit,
        })
    return records

def parse_period_text(text: str) -> tuple[date | None, date | None]:
    """Parse strings like '1/19/13-2/8/13' or '12/4-1/5' into date pair."""
    s = text.strip().replace(" ", "")
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2,4})\s*[-–]\s*(\d{1,2})/(\d{1,2})/(\d{2,4})$", s)
    if not m:
        return None, None
    try:
        m1, d1, y1, m2, d2, y2 = [int(x) for x in m.groups()]
        if y1 < 100: y1 += 2000
        if y2 < 100: y2 += 2000
        if not (2000 <= y1 <= 2099 and 2000 <= y2 <= 2099):
            return None, None
        return date(y1, m1, d1), date(y2, m2, d2)
    except Exception:
        return None, None

def match_property_from_filename(filename: str) -> str | None:
    """
    Map a water breakdown filename to a property code. Order of checks:
      1. Strict explicit mapping (full phrase match)
      2. Give up — unknown property (e.g. Tower Ridge, no longer in portfolio)
    """
    lo = filename.lower().replace("_", " ").replace(".xlsx", "").replace(".xls", "")
    # Strict, explicit, longest-match-wins. Order matters — put discriminating
    # phrases BEFORE shorter/ambiguous ones.
    explicit = [
        ("residences at beverly park",  "611"),
        ("residence at beverly park",   "611"),
        ("naranja lakes",               "610"),
        ("crystal cove",                "608"),
        ("marathon key",                "607"),
        ("haverhill",                   "606"),
        ("heritage at pompano",         "604"),
        ("heritage at prompano",        "604"),    # source typo
        ("pompano",                     "604"),
        ("prompano",                    "604"),
        ("crystal lakes",               "603"),
        ("crystal lake",                "603"),
        ("vista grand",                 "602"),
        ("town park",                   "601"),
        ("ec tyler",                    "562"),
        ("earl campbell",               "562"),
        ("stalcup",                     "561"),
        ("buttercup",                   "561"),
        ("heritage park vista",         "560"),
        ("eastland",                    "559"),
        ("onion creek",                 "558"),
        ("river valley",                "558"),
        ("sunset pointe",               "555"),
        ("heritage at mcdonough",       "516"),
        ("heritage mcdonough",          "516"),
        ("tuscany village",             "515"),
        ("tuscany",                     "515"),
        ("hidden creste",               "514"),
        ("heritage at walton reserve",  "509"),
        ("walton reserve",              "509"),
        ("hearthstone landing",         "508"),
        ("hearthstone",                 "508"),
        ("hl canton",                   "508"),
    ]
    for phrase, code in explicit:
        if phrase in lo:
            return code
    return None

# -----------------------------------------------------------------------------
# SQL emitter
# -----------------------------------------------------------------------------

def emit_sql(
    summary_rows: list[tuple[str, dict]],
    water_history: dict[str, list[dict]],
    water_details: dict[str, dict],
    house_meters: dict[str, dict],
    output_path: Path,
):
    """
    summary_rows:  list of (property_code, row_dict)
    water_history: { property_code: [ record_dict, ... ] }
    water_details: { property_code: { vendor_name, account_number, line_items } }
    house_meters:  { property_code: { vendor_name, meters: [...] } }
    """
    out = []
    out.append("-- ============================================================================")
    out.append("-- Historical utility data — auto-generated by scripts/import-historical.py")
    out.append(f"-- Generated: {datetime.now().isoformat(timespec='seconds')}")
    out.append("-- Apply AFTER 0001_initial_schema, 0002_seed_reference_data, 0003_storage_setup,")
    out.append("-- 0004_sage_batches. Re-runnable — all inserts are idempotent.")
    out.append("-- ============================================================================")
    out.append("")

    # Property 611 — added after the initial seed
    out.append("-- 611 Residences at Beverly Park — added with historical data")
    out.append("insert into properties (code, full_code, name, short_name, state) values")
    out.append("  ('611', '500-611', 'Residences at Beverly Park', 'RBP', 'FL')")
    out.append("on conflict (code) do nothing;")
    out.append("")

    # -------------------------------------------------------------------------
    # Monthly Summary → historical invoices (one per property/GL/month/description)
    # -------------------------------------------------------------------------
    if summary_rows:
        out.append(f"-- {len(summary_rows):,} Summary-sheet rows from {len(set(p for p,_ in summary_rows))} properties")
        out.append("with hist(property_code, year, month, gl_code, description, amount) as (values")
        lines = []
        for pc, r in summary_rows:
            lines.append(
                f"  ({sql_str(pc)}, {r['year']}, {r['month']}, "
                f"{sql_str(r['gl_code'])}, {sql_str(r['description'])}, {r['amount']})"
            )
        out.append(",\n".join(lines))
        out.append(")")
        out.append("""insert into invoices
  (property_id, gl_account_id,
   invoice_number, invoice_date, due_date,
   service_period_start, service_period_end, service_days,
   current_charges, total_amount_due, status, source, source_reference,
   sage_posted_at, gl_coding, extraction_confidence, requires_human_review)
select
  p.id, g.id,
  'HIST-S-' || h.property_code || '-' || h.gl_code || '-'
    || regexp_replace(h.description, '[^A-Za-z0-9]', '', 'g') || '-'
    || h.year || '-' || lpad(h.month::text, 2, '0'),
  make_date(h.year, h.month, 1),
  make_date(h.year, h.month, 28),
  make_date(h.year, h.month, 1),
  (date_trunc('month', make_date(h.year, h.month, 1)) + interval '1 month - 1 day')::date,
  extract(day from (date_trunc('month', make_date(h.year, h.month, 1))
                    + interval '1 month - 1 day'))::int,
  h.amount, h.amount, 'paid', 'manual',
  'historical_import_summary',
  make_date(h.year, h.month, 1),
  '500-' || p.code || '-' || g.code || '.00',
  1.00, false
from hist h
  join properties  p on p.code = h.property_code
  join gl_accounts g on g.code = h.gl_code
on conflict do nothing;""")
        out.append("")

    # -------------------------------------------------------------------------
    # Water meter vendors + accounts (synthetic — real accounts added later)
    # -------------------------------------------------------------------------
    if water_history:
        out.append("-- Synthetic water vendors + utility_accounts for historical usage linkage.")
        out.append("-- These are placeholders; real vendors/accounts are added as bills arrive.")
        out.append("insert into vendors (name, category) values")
        out.append("  ('Water utility (historical)', 'water')")
        out.append("on conflict do nothing;")
        out.append("")

        out.append("-- One historical utility_account per property (for usage_reading linkage)")
        ua_lines = []
        for pc in sorted(water_history.keys()):
            ua_lines.append(
                f"  ({sql_str(pc)}, {sql_str('HIST-' + pc)})"
            )
        out.append("with ua(property_code, account_number) as (values")
        out.append(",\n".join(ua_lines))
        out.append(")")
        out.append("""insert into utility_accounts
  (property_id, vendor_id, gl_account_id, account_number, description, sub_code,
   baseline_window_months, variance_threshold_pct, usage_unit, active)
select
  p.id, v.id, g.id, ua.account_number,
  'Historical water baseline', '00', 12, 3.00, 'gallons', true
from ua
  join properties  p on p.code = ua.property_code
  join vendors     v on v.name = 'Water utility (historical)'
  join gl_accounts g on g.code = '5120'
on conflict (vendor_id, account_number) do nothing;""")
        out.append("")

        # ---------------------------------------------------------------------
        # Water invoices + usage readings
        # ---------------------------------------------------------------------
        total_readings = sum(len(v) for v in water_history.values())
        out.append(f"-- {total_readings:,} historical water readings from {len(water_history)} properties")
        out.append("with wr(property_code, period_start, period_end, days, usage, daily, occupancy, unit) as (values")
        lines = []
        for pc in sorted(water_history.keys()):
            for rec in water_history[pc]:
                lines.append(
                    f"  ({sql_str(pc)}, "
                    f"{sql_str(rec['period_start'])}, {sql_str(rec['period_end'])}, "
                    f"{rec['days'] if rec['days'] else 'null'}, "
                    f"{rec['usage'] if rec['usage'] else 'null'}, "
                    f"{rec['daily_usage'] if rec['daily_usage'] else 'null'}, "
                    f"{rec['occupancy'] if rec['occupancy'] else 'null'}, "
                    f"{sql_str(rec.get('unit', 'gallons'))})"
                )
        out.append(",\n".join(lines))
        out.append(")")
        # Insert invoices
        out.append(""", invs as (
  insert into invoices
    (property_id, vendor_id, utility_account_id, gl_account_id,
     invoice_number, invoice_date, due_date,
     service_period_start, service_period_end, service_days,
     current_charges, total_amount_due, status, source, source_reference,
     sage_posted_at, gl_coding, extraction_confidence, requires_human_review)
  select
    p.id, ua.vendor_id, ua.id, g.id,
    'HIST-W-' || wr.property_code || '-' || to_char(wr.period_start::date, 'YYYYMMDD'),
    wr.period_start::date, wr.period_end::date, wr.period_start::date, wr.period_end::date,
    wr.days,
    0, 0, 'paid', 'manual',
    'historical_import_water_usage',
    wr.period_end::date,
    '500-' || p.code || '-5120.00',
    1.00, false
  from wr
    join properties p on p.code = wr.property_code
    join utility_accounts ua on ua.account_number = ('HIST-' || wr.property_code)
    join gl_accounts g on g.code = '5120'
  on conflict do nothing
  returning id, utility_account_id, invoice_number, service_period_start, service_period_end, service_days
)
insert into usage_readings
  (invoice_id, utility_account_id, reading_type,
   service_start, service_end, days,
   usage_amount, usage_unit, occupancy_pct)
select
  i.id, i.utility_account_id, 'water',
  i.service_period_start, i.service_period_end, i.service_days,
  wr.usage, wr.unit, wr.occupancy
from invs i
  join wr on ('HIST-W-' || wr.property_code || '-' ||
              to_char(wr.period_start::date, 'YYYYMMDD')) = i.invoice_number
where wr.usage is not null;""")
        out.append("")

    # -------------------------------------------------------------------------
    # Water-detail line items
    # -------------------------------------------------------------------------
    # For each property that had a Water sheet parsed, create invoice_line_items
    # rows linked to the monthly HIST-S-<code>-5120-Water-<year>-<mm> invoices
    # that the Summary block created above. Each month typically gets 3-5 lines
    # (water, sewer, irrigation, storm water, envir protection).
    #
    # This runs AFTER the summary block since it needs the invoices to exist.

    if water_details:
        total_line_items = sum(len(d["line_items"]) for d in water_details.values())
        out.append(f"-- {total_line_items:,} water-detail line items across {len(water_details)} properties")
        out.append("-- Linked to the HIST-S-<code>-<gl>-Water-<year>-<mm> invoice rows created above.")
        out.append("")

        out.append("with li(property_code, year, month, description, amount, category, gl_code, is_consumption_based) as (values")
        lines = []
        for pc in sorted(water_details.keys()):
            for item in water_details[pc]["line_items"]:
                lines.append(
                    f"  ({sql_str(pc)}, {item['year']}, {item['month']}, "
                    f"{sql_str(item['description'])}, {item['amount']}, "
                    f"{sql_str(item['category'])}, {sql_str(item['gl_code'])}, "
                    f"{'true' if item['is_consumption_based'] else 'false'})"
                )
        out.append(",\n".join(lines))
        out.append(")")
        # Match each line-item back to its parent summary invoice by property/gl/month
        out.append("""insert into invoice_line_items
  (invoice_id, gl_account_id, sub_code, gl_coding,
   description, category, amount, is_consumption_based, source_row_label)
select
  i.id, g.id, '00',
  '500-' || p.code || '-' || g.code || '.00',
  li.description, li.category, li.amount, li.is_consumption_based,
  li.description
from li
  join properties p   on p.code = li.property_code
  join gl_accounts g  on g.code = li.gl_code
  join invoices i     on i.property_id = p.id
                     and i.source_reference = 'historical_import_summary'
                     and extract(year  from i.invoice_date)::int = li.year
                     and extract(month from i.invoice_date)::int = li.month
                     -- Match to the primary water summary row for that month
                     and (i.invoice_number like 'HIST-S-' || p.code || '-5120-Water-%'
                       or i.invoice_number like 'HIST-S-' || p.code || '-5120-Storm%'
                       or i.invoice_number like 'HIST-S-' || p.code || '-5120-Envir%'
                       or i.invoice_number like 'HIST-S-' || p.code || '-5120-Cluhouse%'
                       or i.invoice_number like 'HIST-S-' || p.code || '-5120-Water')
                     and i.gl_account_id = (select id from gl_accounts where code = '5120')
on conflict do nothing;""")
        out.append("")

        # Reconciliation: mark invoices where sum of line items equals total
        out.append("-- Reconciliation flag — true where line-items sum to invoice total")
        out.append("""update invoices i
set line_items_reconciled = case
  when li_totals.line_items_total is null then null
  when abs(coalesce(li_totals.line_items_total, 0) - i.total_amount_due) < 0.02 then true
  else false
end
from v_invoice_line_totals li_totals
where li_totals.invoice_id = i.id
  and i.source_reference = 'historical_import_summary';""")
        out.append("")

    # -------------------------------------------------------------------------
    # House Meters — per-meter utility accounts + monthly invoices
    # -------------------------------------------------------------------------
    # For each property's House Meters sheet:
    #   1. Register the electric vendor
    #   2. Create one utility_accounts row per meter (with meter_id/esi_id/
    #      meter_category populated)
    #   3. Create one HIST-M-<code>-<meter>-<yyyymm> invoice per meter per
    #      month with data

    if house_meters:
        total_meters = sum(len(d["meters"]) for d in house_meters.values())
        out.append(f"-- {total_meters} historical house meters across {len(house_meters)} properties")
        out.append("")

        # Register each electric vendor encountered
        vendors_seen = set()
        for pc, data in sorted(house_meters.items()):
            v = data.get("vendor_name")
            if v and v not in vendors_seen:
                vendors_seen.add(v)
        if vendors_seen:
            out.append("-- Electric vendors discovered in historical House Meters data")
            for v in sorted(vendors_seen):
                out.append(f"insert into vendors (name, category) values ({sql_str(v)}, 'electric')")
                out.append("on conflict do nothing;")
            out.append("")

        # Upsert utility_accounts rows — one per (property × meter)
        out.append("-- One utility_accounts row per physical meter")
        out.append("with meters_src(property_code, vendor_name, account_number, meter_id, esi_id,")
        out.append("                 description, category, gl_code) as (values")
        meter_rows = []
        for pc in sorted(house_meters.keys()):
            data = house_meters[pc]
            vendor = data.get("vendor_name") or "Electric utility (historical)"
            for m in data["meters"]:
                # Account number fallback — use meter_id or synthetic if both null
                acct = m.get("account_number") or m.get("meter_id") or m.get("esi_id")
                if not acct:
                    acct = f"HIST-M-{pc}-{abs(hash(m['description'])) % 100000}"
                meter_rows.append(
                    f"  ({sql_str(pc)}, {sql_str(vendor)}, {sql_str(acct)}, "
                    f"{sql_str(m.get('meter_id'))}, {sql_str(m.get('esi_id'))}, "
                    f"{sql_str(m.get('description'))}, {sql_str(m.get('category'))}, "
                    f"{sql_str(m.get('gl_code'))})"
                )
        out.append(",\n".join(meter_rows))
        out.append(")")
        # First ensure the vendors exist (second insert, idempotent)
        out.append("""
, ensure_vendor as (
  insert into vendors (name, category)
  select distinct vendor_name, 'electric' from meters_src
  on conflict (name) do nothing
  returning id, name
)
insert into utility_accounts
  (property_id, vendor_id, gl_account_id, account_number,
   meter_id, esi_id, meter_category,
   description, sub_code,
   baseline_window_months, variance_threshold_pct, usage_unit, active)
select
  p.id, v.id, g.id, m.account_number,
  m.meter_id, m.esi_id, m.category,
  coalesce(m.description, 'Meter ' || m.meter_id),
  '00', 12, 3.00, 'kwh', true
from meters_src m
  join properties  p on p.code = m.property_code
  join vendors     v on v.name = m.vendor_name
  join gl_accounts g on g.code = m.gl_code
on conflict (vendor_id, account_number) do update
  set meter_id        = excluded.meter_id,
      esi_id          = excluded.esi_id,
      meter_category  = excluded.meter_category,
      description     = excluded.description;""")
        out.append("")

        # Emit historical invoices — one per meter × month
        out.append("-- Historical monthly invoices per meter (status='paid')")
        out.append("with meter_months(property_code, account_number, year, month, amount) as (values")
        mm_rows = []
        for pc in sorted(house_meters.keys()):
            data = house_meters[pc]
            year = data.get("year", date.today().year)
            for m in data["meters"]:
                acct = m.get("account_number") or m.get("meter_id") or m.get("esi_id")
                if not acct:
                    acct = f"HIST-M-{pc}-{abs(hash(m['description'])) % 100000}"
                for month, amount in m["monthly_amounts"].items():
                    mm_rows.append(
                        f"  ({sql_str(pc)}, {sql_str(acct)}, {year}, {month}, {amount})"
                    )
        if mm_rows:
            out.append(",\n".join(mm_rows))
            out.append(")")
            out.append("""insert into invoices
  (property_id, vendor_id, utility_account_id, gl_account_id,
   invoice_number, invoice_date, due_date,
   service_period_start, service_period_end, service_days,
   current_charges, total_amount_due, status, source, source_reference,
   sage_posted_at, gl_coding, extraction_confidence, requires_human_review)
select
  p.id, ua.vendor_id, ua.id, g.id,
  'HIST-M-' || mm.property_code || '-' || regexp_replace(mm.account_number, '[^A-Za-z0-9]', '', 'g')
    || '-' || mm.year || lpad(mm.month::text, 2, '0'),
  make_date(mm.year, mm.month, 1),
  make_date(mm.year, mm.month, 28),
  make_date(mm.year, mm.month, 1),
  (date_trunc('month', make_date(mm.year, mm.month, 1)) + interval '1 month - 1 day')::date,
  extract(day from (date_trunc('month', make_date(mm.year, mm.month, 1))
                    + interval '1 month - 1 day'))::int,
  mm.amount, mm.amount, 'paid', 'manual',
  'historical_import_meter',
  make_date(mm.year, mm.month, 1),
  '500-' || p.code || '-' || ua.gl_account_id::text || '.00',
  1.00, false
from meter_months mm
  join properties p on p.code = mm.property_code
  join utility_accounts ua
    on ua.property_id = p.id and ua.account_number = mm.account_number
  join gl_accounts g on g.id = ua.gl_account_id
on conflict do nothing;""")
            out.append("")

        # Clean up the Summary-sheet "House Electric" rows that are superseded
        # by the per-meter detail above
        out.append("-- Remove Summary-sheet House Electric rows superseded by per-meter detail")
        out.append("""delete from invoices
  where source_reference = 'historical_import_summary'
    and gl_account_id in (select id from gl_accounts where code in ('5112', '5114', '5116'))
    and property_id in (
      select distinct property_id from invoices
      where source_reference = 'historical_import_meter'
    );""")
        out.append("")

    out.append("-- End of historical import")
    output_path.write_text("\n".join(out))

# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------

def main():
    p = argparse.ArgumentParser()
    p.add_argument("--input",  required=True)
    p.add_argument("--output", required=True)
    args = p.parse_args()

    input_dir = Path(args.input)
    output    = Path(args.output)
    workdir   = Path(tempfile.mkdtemp(prefix="nurock_hist_"))

    # Only process .xls files in input dir (skip any pre-converted .xlsx duplicates)
    files = [p for p in sorted(input_dir.iterdir())
             if p.is_file() and p.suffix.lower() in (".xls", ".xlsx")]

    # De-dupe: prefer .xls (original) over .xlsx (converted copy)
    by_stem = {}
    for f in files:
        stem = f.stem
        # If we have both .xls and .xlsx for the same stem, prefer .xls
        if stem not in by_stem or f.suffix.lower() == ".xls":
            if stem not in by_stem or f.suffix.lower() == ".xls":
                by_stem[stem] = f
    files = sorted(by_stem.values())

    summary_rows  = []    # [(property_code, row_dict)]
    water_history = {}    # {property_code: [record_dict]}
    water_details = {}    # {property_code: {vendor_name, account_number, line_items}}
    house_meters  = {}    # {property_code: {vendor_name, meters: [...]}}

    for path in files:
        name = path.name
        print(f"→ {name}")
        xlsx = ensure_xlsx(path, workdir)
        if not xlsx:
            print("  ✗ conversion failed")
            continue

        try:
            wb = load_workbook(xlsx, data_only=True)
        except Exception as e:
            print(f"  ✗ load failed: {e}")
            continue

        # Is this a per-property workbook or a water-usage breakdown?
        pc = property_code_from_filename(name)
        is_breakdown = ("break" in name.lower() or
                        "usage" in name.lower() or
                        "consumption" in name.lower())

        if pc and pc in PROPERTY_BY_CODE and not is_breakdown:
            rows = parse_summary(wb, pc)
            for r in rows:
                summary_rows.append((pc, r))
            print(f"  ✓ Summary: {len(rows)} monthly rows")

            # Also parse Water sheet for line-item detail
            detail = parse_water_detail(wb, pc, date.today().year)
            if detail and detail.get("line_items"):
                water_details[pc] = detail
                months_with_data = set(li["month"] for li in detail["line_items"])
                print(f"    + Water detail: {len(detail['line_items'])} line items "
                      f"across {len(months_with_data)} months "
                      f"(vendor={detail.get('vendor_name')!r})")

            # And the House Meters sheet for per-meter detail
            hm = parse_house_meters(wb, pc, date.today().year)
            if hm and hm.get("meters"):
                house_meters[pc] = hm
                total_dollars = sum(
                    sum(m["monthly_amounts"].values()) for m in hm["meters"]
                )
                print(f"    + House meters: {len(hm['meters'])} meters, "
                      f"${float(total_dollars):,.0f} annual "
                      f"(vendor={hm.get('vendor_name')!r})")
        elif is_breakdown:
            bp = match_property_from_filename(name)
            if not bp:
                print("  ✗ could not match to a known property")
                continue
            recs = parse_water_breakdown(wb)
            if recs:
                water_history.setdefault(bp, []).extend(recs)
                yr_min = min(r["period_start"].year for r in recs if r["period_start"])
                yr_max = max(r["period_end"].year for r in recs if r["period_end"])
                print(f"  ✓ Water breakdown for {bp}: {len(recs)} readings ({yr_min}–{yr_max})")
            else:
                print(f"  ⚠ no readings parsed for {bp}")

    print()
    total_summary = len(summary_rows)
    total_water   = sum(len(v) for v in water_history.values())
    total_lines   = sum(len(d["line_items"]) for d in water_details.values())
    total_meters  = sum(len(d["meters"]) for d in house_meters.values())
    print(f"Summary rows:         {total_summary:,} across {len(set(pc for pc,_ in summary_rows))} properties")
    print(f"Water readings:       {total_water:,} across {len(water_history)} properties")
    print(f"Water line items:     {total_lines:,} across {len(water_details)} properties")
    print(f"House meters:         {total_meters:,} across {len(house_meters)} properties")

    emit_sql(summary_rows, water_history, water_details, house_meters, output)
    print(f"\nWrote {output} ({output.stat().st_size:,} bytes)")

if __name__ == "__main__":
    main()
